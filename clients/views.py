
from decimal import Decimal
from django.db import transaction
    
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password
from django.views.decorators.http import require_POST, require_GET, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.middleware.csrf import get_token
from django.utils import timezone
from django.http import (
    HttpResponse,
    JsonResponse,
    HttpResponseBadRequest,
    Http404
)
from django.db.models import (
    Sum, Count, Value, F, DecimalField, Q, Prefetch
)
from django.db.models.functions import Coalesce
from django.forms.models import model_to_dict

from .models import (
    Client,
    Loan,
    LoanRepaymentDay,
    Announcement,
    TermsAndConditions,
    DailyReport
)
from .forms import (
    ClientEditForm,
    TermsAndConditionsForm,
    DailyReportForm,
    LoanOfficerCreationForm
)
from clients.models import Client, Loan

from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
import io
import json
from django.contrib.auth import get_user_model
from .forms import LoanOfficerEditProfileForm
User = get_user_model()



from .forms import (
    ClientForm,
    DailyReportForm,
    AnnouncementForm,
    TermsAndConditionsForm,
    CEOEditProfileForm as CEOProfileForm,
)
from .models import (
    Client,
    Announcement,
    DailyReport,
    TermsAndConditions,
)

# Additional imports
import io
from datetime import datetime, date
from reportlab.pdfgen import canvas

@login_required
def register_client(request):
    if request.user.role != 'OFFICER':
        return render(request, '403.html')

    if request.method == 'POST':
        form = ClientForm(request.POST, request.FILES)
        if form.is_valid():
            client = form.save(commit=False)
            client.created_by = request.user
            client.approved = False
            client.save()
            return redirect('officer_dashboard')
    else:
        form = ClientForm()

    return render(request, 'clients/register_client.html', {'form': form})


@login_required
def view_my_clients(request):
    clients = Client.objects.filter(
        created_by=request.user
    ).order_by('-created_at')

    return render(request, 'clients/view_my_clients.html', {'clients': clients})


@login_required
def submit_daily_report(request):
    today = timezone.localdate()
    user = request.user


    approved_clients = Client.objects.filter(created_by=user, approved=True)
    total_expected = Decimal('0.00')
    clients_owing = 0

    # Calculate expected amount for today

    approved_clients = Client.objects.filter(created_by=user, approved=True)
    total_expected = Decimal('0.00')
    total_collected = Decimal('0.00')
    clients_owing = 0


    for client in approved_clients:
        loan = client.loans.filter(approved=True).order_by('-created_at').first()
        if not loan or loan.created_at.date() == today:
            continue
        if today.strftime('%A').upper() == loan.exempt_day.upper():
            continue

        today_day = loan.repayment_days.filter(date=today).first()
        if not today_day:
            continue

        today_day = loan.repayment_days.filter(date=today).first()
        if not today_day:
            continue


        previous_paid_day = loan.repayment_days.filter(date__lt=today, is_paid=True).order_by('-date').first()
        carry_forward = Decimal('0.00')
        if previous_paid_day:
            carry_forward = (previous_paid_day.amount_due or 0) - (previous_paid_day.amount_paid or 0)

        effective_due_today = (today_day.amount_due or 0) + carry_forward
        total_expected += effective_due_today
        if not today_day.is_paid:
            clients_owing += 1

    # Sum all payments marked today (including advance payments for future days)
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today, datetime.max.time())
    repayments_today = LoanRepaymentDay.objects.filter(
        loan__client__created_by=user,
        marked_at__date=today,
        is_paid=True
    )
    total_collected = repayments_today.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

    # Advance payments: sum of payments for future days marked today
    advance_payments = repayments_today.filter(date__gt=today).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

    # Balance: expected - collected (if negative, set to 0)
    balance = total_expected - (total_collected - advance_payments)
    if balance < 0:


        effective_due_today = (today_day.amount_due or 0) + carry_forward
        total_expected += effective_due_today

        if today_day.is_paid:
            total_collected += today_day.amount_paid or 0
        else:
            clients_owing += 1

    if total_collected > total_expected:
        advance_payments = total_collected - total_expected
        balance = Decimal('0.00')
    elif total_collected < total_expected:
        balance = total_expected - total_collected
        advance_payments = Decimal('0.00')
    else:
        advance_payments = Decimal('0.00')
        balance = Decimal('0.00')

    accumulative_collected = LoanRepaymentDay.objects.filter(
        loan__client__created_by=user,
        is_paid=True
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

    report_instance = DailyReport.objects.filter(submitted_by=user, date=today).first()

    if request.method == 'POST':
        optional_note = request.POST.get('optional_note', '').strip()


        if report_instance:
            report = report_instance
        else:
            report = DailyReport(submitted_by=user, date=today)

        report.total_expected = total_expected
        report.total_collected = total_collected
        report.advance_payments = advance_payments
        report.balance = balance
        report.accumulative_balance = accumulative_collected
        report.clients_owing = clients_owing
        report.optional_note = optional_note
        report.is_submitted = True
        report.save()

        messages.success(request, " Daily report submitted.")
        return redirect('submit_daily_report')


        messages.success(request, " Daily report submitted.")
        return redirect('submit_daily_report')


    initial_note = ''
    if report_instance and not report_instance.is_submitted:
        initial_note = report_instance.optional_note

    form = DailyReportForm(initial={
        'date': today,
        'total_expected': total_expected,
        'total_collected': total_collected,
        'advance_payments': advance_payments,
        'balance': balance,
        'accumulative_balance': accumulative_collected,
        'clients_owing': clients_owing,
        'optional_note': initial_note,
    })

    return render(request, 'clients/submit_daily_report.html', {'form': form})


@login_required
def view_approved_clients(request):
    approved_clients = Client.objects.filter(
        created_by=request.user,
        approved=True
    ).distinct().order_by('-created_at')

    for client in approved_clients:
        latest_loan = client.loans.filter(approved=True).order_by('-created_at').first()

        client.latest_loan = latest_loan

        if latest_loan:
            due_date = latest_loan.created_at.date() + timedelta(days=latest_loan.repayment_days.count())
            client.due_date = due_date

            has_unpaid = latest_loan.repayment_days.filter(is_paid=False).exists()
            client.can_apply = not has_unpaid
        else:
            client.can_apply = True

    return render(request, 'clients/view_approved_clients.html', {
        'approved_clients': approved_clients
    })



from django.db.models import Sum

@login_required
def print_monthly_summary(request):
    today = timezone.now()
    current_month = today.month
    current_year = today.year

    approved_clients = Client.objects.filter(
        created_by=request.user,
        approved=True,
        created_at__month=current_month,
        created_at__year=current_year
    )

    total_amount = Decimal('0.00')

    for client in approved_clients:
        latest_loan = client.loans.filter(approved=True).order_by('-created_at').first()

        if latest_loan:
            client.amount = latest_loan.amount
            client.amount_with_interest = round(latest_loan.amount * Decimal('1.378'), 2)
            total_amount += client.amount_with_interest

            # Sum actual payments instead of estimated
            total_paid = latest_loan.repayment_days.filter(is_paid=True)\
                .aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

            client.total_paid = round(total_paid, 2)
            client.balance = round(client.amount_with_interest - total_paid, 2)
        else:
            client.amount = Decimal('0.00')
            client.amount_with_interest = Decimal('0.00')
            client.total_paid = Decimal('0.00')
            client.balance = Decimal('0.00')

    return render(request, 'clients/print_monthly_summary.html', {
        'clients': approved_clients,
        'total_clients': approved_clients.count(),
        'total_amount': round(total_amount, 2),
        'month': today.strftime('%B'),
        'year': today.year
    })


def is_ceo(user):
    return user.groups.filter(name='CEO').exists()


def view_pending_clients(request):
    loans = Loan.objects.filter(approved=False).select_related('client')
    
    pending_clients = [
        {
            'client': loan.client,
            'loan': loan
        }
        for loan in loans
    ]

    return render(request, 'clients/ceo/view_pending_clients.html', {'pending_clients': pending_clients})


def generate_repayment_schedule(start_date, exempt_day='SUNDAY'):
    dates = []
    current_date = start_date
    while len(dates) < 26:
        if exempt_day.upper() == 'SUNDAY' and current_date.weekday() != 6:
            dates.append(current_date)
        elif exempt_day.upper() == 'SATURDAY' and current_date.weekday() != 5:
            dates.append(current_date)
        current_date += timedelta(days=1)
    return dates


def approve_loan(request, loan_id):
    if request.method == 'POST':
        loan = get_object_or_404(Loan, id=loan_id)

        if loan.approved:
            messages.info(request, "Loan is already approved.")
        else:
            loan.approved = True
            loan.save()
            messages.success(request, f"Loan for {loan.client.full_name} has been approved.")

    return redirect('view_pending_clients')


def reject_loan(request, loan_id):
    if request.method == 'POST':
        loan = get_object_or_404(Loan, id=loan_id)
        loan.delete()
    return redirect('view_pending_clients')


def list_announcements(request):
    announcements = Announcement.objects.order_by('-created_at')
    return render(request, 'clients/ceo_announcements.html', {'announcements': announcements})


def create_announcement(request):
    if request.method == 'POST':
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            announcement = form.save(commit=False)
            announcement.created_by = request.user
            announcement.save()
            return redirect('list_announcements')
    else:
        form = AnnouncementForm()
    return render(request, 'clients/create_announcement.html', {'form': form})


def ceo_announcements(request):
    announcements = Announcement.objects.all()
    return render(request, "clients/ceo_announcements.html", {"announcements": announcements})


def delete_announcement(request, id):
    if request.method == 'POST':
        announcement = get_object_or_404(Announcement, id=id)
        announcement.delete()
        messages.success(request, "Announcement deleted successfully.")
    return redirect('list_announcements')


@login_required
def ceo_reports_analytics(request):

    total_payments = LoanRepaymentDay.objects.filter(is_paid=True).count()

    total_amount = LoanRepaymentDay.objects.filter(is_paid=True).aggregate(
        total=Sum('amount_paid')
    )['total'] or 0

    payments_by_officer = LoanRepaymentDay.objects.filter(is_paid=True).values(
        'marked_by__username'
    ).annotate(
        total=Count('id'),
        amount=Sum('amount_paid')
    ).order_by('-amount')

    context = {
        'total_payments': total_payments,
        'total_amount': total_amount,
        'payments_by_officer': payments_by_officer,
    }

    return render(request, 'clients/ceo_reports_analytics.html', context)

@login_required
def ceo_reports_analytics(request):
    repayment_days = LoanRepaymentDay.objects.select_related('loan__client')
    query_date = request.GET.get('date')
    query_month = request.GET.get('month')

    if query_date:
        repayment_days = repayment_days.filter(date=query_date)
    elif query_month:
        repayment_days = repayment_days.filter(date__month=query_month)

    paid_days = repayment_days.filter(is_paid=True)

    total_amount = paid_days.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

    if 'export' in request.GET:
        export_type = request.GET['export']
        if export_type == 'pdf':
            buffer = io.BytesIO()
            p = canvas.Canvas(buffer)
            p.setFont("Helvetica", 12)
            y = 800
            p.drawString(100, y, f"Loan Repayment Report - {datetime.now().date()}")
            y -= 30
            for day in paid_days:
                paid_amount = day.amount_paid or Decimal('0.00')
                p.drawString(
                    100, y,
                    f"{day.loan.client.full_name} - ZMW {paid_amount:.2f} on {day.date.strftime('%Y-%m-%d')}"
                )
                y -= 20
                if y < 50:
                    p.showPage()
                    y = 800
            p.drawString(100, y-20, f"Total Collected: ZMW {total_amount:.2f}")
            p.showPage()
            p.save()
            buffer.seek(0)
            return HttpResponse(buffer, content_type='application/pdf')
        elif export_type == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Repayments"
            # Header
            ws.append(["Client", "Amount (ZMW)", "Date"])
            for cell in ws[1]:
                cell.font = Font(bold=True)
            # Data
            for day in paid_days:
                ws.append([
                    day.loan.client.full_name,
                    float(day.amount_paid or 0.0),
                    day.date.strftime('%Y-%m-%d')
                ])
            # Total row
            ws.append(["Total", float(total_amount), ""])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=repayments_report.xlsx'
            wb.save(response)
            return response

    if 'export' in request.GET and request.GET['export'] == 'pdf':
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer)
        p.setFont("Helvetica", 12)
        y = 800
        p.drawString(100, y, f"Loan Repayment Report - {datetime.now().date()}")
        y -= 30
        for day in paid_days:
            paid_amount = day.amount_paid or Decimal('0.00')
            p.drawString(
                100, y,
                f"{day.loan.client.full_name} - ZMW {paid_amount:.2f} on {day.date.strftime('%Y-%m-%d')}"
            )
            y -= 20
            if y < 50:
                p.showPage()
                y = 800
        p.drawString(100, y-20, f"Total Collected: ZMW {total_amount:.2f}")
        p.showPage()
        p.save()
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')

    return render(request, 'clients/ceo_reports_analytics.html', {
        'repayment_days': paid_days,
        'total_amount': round(total_amount, 2),
    })



from .forms import ClientFilterForm
from django.contrib.auth import get_user_model

User = get_user_model()

@login_required
def view_all_clients(request):
    form = ClientFilterForm(request.GET or None)
    clients = Client.objects.all().order_by('-created_at')


    # Officer filter
    if form.is_valid():
        officer_id = form.cleaned_data.get('officer')
        if officer_id:
            clients = clients.filter(created_by__id=officer_id)

    # Status filter
    status = request.GET.get('status')
    if status == 'pending':
        clients = clients.filter(approved=False, rejected=False)
    elif status == 'approved':
        clients = clients.filter(approved=True)

    officers = User.objects.filter(role='OFFICER').order_by('username')

    context = {
        'clients': clients,
        'form': form,
        'officers': officers,
    }
    return render(request, 'clients/view_all_clients.html', context)
    

from django.http import HttpResponse
import csv
from django.utils import timezone

@login_required
def view_all_payments(request):
    selected_date = request.GET.get('date') or timezone.localdate()

    paid_days = LoanRepaymentDay.objects.filter(
        is_paid=True,
        marked_at__date=selected_date
    ).select_related('loan__client')

    payment_records = []
    for day in paid_days:
        loan = day.loan
        actual_paid = day.amount_paid or day.amount_due

        payment_records.append({
            'client_name': loan.client.full_name,
            'nrc': loan.client.nrc,
            'phone_number': loan.client.phone_number,
            'amount_paid': actual_paid,
            'payment_date': day.marked_at or day.date,
            'status': "Paid",
            'id': day.id,
            'edit_note': day.edit_note
        })

    if 'export' in request.GET and request.GET['export'] == 'excel':
        response = HttpResponse(content_type='text/csv')
        filename = f"payments_{selected_date}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(['Client Name', 'NRC', 'Phone', 'Amount Paid', 'Date', 'Status'])
        for record in payment_records:
            writer.writerow([
                record['client_name'],
                record['nrc'],
                record['phone_number'],
                f"{record['amount_paid']:.2f}",
                record['payment_date'].strftime('%Y-%m-%d'),
                record['status']
            ])

        return response

    return render(request, 'clients/ceo_view_all_payments.html', {
        'payments': payment_records,
        'selected_date': selected_date
    })


@login_required
def view_officer_reports(request):
    selected_date = request.GET.get('date')
    
    if not selected_date:
        selected_date = timezone.localdate().isoformat()  # Default to today

    reports = DailyReport.objects.filter(
        is_submitted=True,
        date=selected_date
    ).select_related('submitted_by').order_by('-submitted_at')

    return render(request, 'clients/ceo_officer_reports.html', {
        'reports': reports,
        'selected_date': selected_date
    })


@login_required
def export_reports_xlsx(request):
    reports = DailyReport.objects.filter(is_submitted=True).select_related('submitted_by').order_by('-submitted_at')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Officer Reports"


    headers = [
        "Officer", "Date", "Expected (ZMW)", "Collected (ZMW)",
        "Advance (ZMW)", "Balance", "Accumulative", "Clients Owing", "Note"
    ]
    sheet.append(headers)

    for report in reports:
        sheet.append([
            report.submitted_by.username,
            report.date.strftime('%Y-%m-%d'),
            float(report.total_expected),
            float(report.total_collected),
            float(report.advance_payments),
            float(report.balance),
            float(report.accumulative_balance),
            report.clients_owing,
            report.optional_note or ""
        ])

    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=officer_reports.xlsx'
    workbook.save(response)
    return response


@login_required
def expected_daily_amounts(request):
    User = get_user_model()
    officers = User.objects.filter(role='OFFICER')
    today = timezone.localdate()
    data = []

    for officer in officers:
        clients = Client.objects.filter(created_by=officer, approved=True)
        total_due_today = Decimal('0.00')
        clients_paid_today = 0
        client_count = 0

        for client in clients:
            loan = client.loans.filter(approved=True).order_by('-created_at').first()
            if not loan:
                continue

            # Skip today's loan
            if loan.created_at.date() == today:
                continue

            # Skip exempt days
            if today.strftime('%A').upper() == loan.exempt_day.upper():
                continue

            repayment_day = loan.repayment_days.filter(date=today).first()
            if not repayment_day:
                continue

            previous_paid_day = loan.repayment_days.filter(
                date__lt=today, is_paid=True
            ).order_by('-date').first()

            carry_forward = Decimal('0.00')
            if previous_paid_day:
                due_yesterday = previous_paid_day.amount_due or Decimal('0.00')
                paid_yesterday = previous_paid_day.amount_paid or Decimal('0.00')
                carry_forward = due_yesterday - paid_yesterday  # Shortfall is positive, excess is negative

            effective_due = (repayment_day.amount_due or Decimal('0.00')) + carry_forward
            total_due_today += effective_due
            client_count += 1

            if repayment_day.is_paid:
                clients_paid_today += 1

        data.append({
            'officer': officer,
            'expected_today': round(total_due_today, 2),
            'client_count': client_count,
            'clients_paid_today': clients_paid_today,
        })

    return render(request, 'ceo_expected_daily_amounts.html', {'officer_data': data})


@login_required
def ceo_accumulated_balances(request):
    clients = Client.objects.filter(loans__approved=True).distinct()

    client_data = []

    for client in clients:
        approved_loans = client.loans.filter(approved=True)

        total_loan_amount = Decimal('0.00')
        total_paid = Decimal('0.00')

        for loan in approved_loans:
            amount_with_interest = loan.total_with_interest
            total_loan_amount += amount_with_interest

            # Sum actual payments from repayment days
            paid_sum = loan.repayment_days.filter(is_paid=True).aggregate(
                total=Sum('amount_paid')
            )['total'] or Decimal('0.00')

            total_paid += paid_sum

        balance = total_loan_amount - total_paid

        client_data.append({
            'client': client,
            'total_loan_amount': round(total_loan_amount, 2),
            'total_paid': round(total_paid, 2),
            'balance': round(balance, 2),
        })

    return render(request, 'clients/ceo_accumulated_balances.html', {
        'clients_data': client_data
    })


def ceo_total_expected_funds(request):
    User = get_user_model()
    officers = User.objects.filter(role='OFFICER')

    officer_data = []
    total_expected = Decimal('0.00')

    for officer in officers:
        approved_clients = Client.objects.filter(created_by=officer, approved=True)

        loans = Loan.objects.filter(client__in=approved_clients)

        total_loans = sum((loan.total_with_interest for loan in loans), Decimal('0.00'))
        total_paid = sum((loan.total_paid() for loan in loans), Decimal('0.00'))
        balance = total_loans - total_paid
        total_expected += balance

        officer_data.append({
            'name': officer.get_full_name() or officer.username,
            'approved_clients': approved_clients.count(),
            'total_loans': total_loans,
            'total_paid': total_paid,
            'balance': balance
        })

    return render(request, 'clients/ceo_total_expected_funds.html', {
        'officer_data': officer_data,
        'total_expected': total_expected
    })

@login_required
def ceo_terms_and_conditions(request):
    all_terms = TermsAndConditions.objects.order_by('-created_at')  # for listing
    form = TermsAndConditionsForm()

    if request.method == 'POST':
        form = TermsAndConditionsForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('ceo_terms_and_conditions')

    return render(request, 'ceo_terms_and_conditions.html', {
        'form': form,
        'terms_list': all_terms
    })


@login_required
def delete_terms_and_conditions(request, pk):
    term = get_object_or_404(TermsAndConditions, pk=pk)
    term.delete()
    return redirect('ceo_terms_and_conditions')



@login_required
def ceo_edit_profile(request):
    user = request.user
    if request.method == 'POST':
        form = CEOProfileForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save(commit=False)
            password = form.cleaned_data.get('password')
            if password:
                user.password = make_password(password)
            user.save()
            return redirect('ceo_dashboard')
    else:
        form = CEOProfileForm(instance=user)
    return render(request, 'ceo_edit_profile.html', {'form': form})


def home(request):
    announcements = Announcement.objects.order_by('-created_at')[:5]  # latest 5
    try:
        terms = TermsAndConditions.objects.latest('last_updated')
    except TermsAndConditions.DoesNotExist:
        terms = None
    return render(request, 'home.html', {
        'announcements': announcements,
        'terms': terms
    })


def client_details_api(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    loans = Loan.objects.filter(client=client).order_by('-created_at')

    loan_data = []

    for loan in loans:
        total_repayable = float(loan.amount) * 1.378

        # Use actual sum of amount_paid, not estimated
        total_paid = loan.repayment_days.filter(is_paid=True)\
            .aggregate(total=Sum('amount_paid'))['total'] or 0

        last_day = loan.repayment_days.order_by('-date').first()
        due_date = last_day.date if last_day else loan.created_at.date() + timedelta(days=35)

        balance = total_repayable - float(total_paid)

        loan_data.append({
            'id': loan.id,
            'original_amount': float(loan.amount),
            'amount': round(total_repayable, 2),
            'approved': loan.approved,
            'created_at': loan.created_at.strftime('%Y-%m-%d'),
            'due_date': due_date.strftime('%Y-%m-%d'),
            'total_paid': round(float(total_paid), 2),
            'balance': round(balance, 2),
            'is_settled': balance <= 0,
        })

    data = {
        'full_name': client.full_name,
        'nrc': client.nrc,
        'phone_number': client.phone_number,
        'created_at': client.created_at.strftime('%Y-%m-%d'),
        'loans': loan_data
    }

    return JsonResponse(data)


@require_POST
@login_required
def apply_loan(request):
    client_id = request.POST.get("client_id")
    amount = request.POST.get("amount")
    exempt_day = request.POST.get("exempt_day")  # "SATURDAY" or "SUNDAY"

    client = get_object_or_404(Client, id=client_id)

    try:
        amount = Decimal(amount)

        loan = Loan.objects.create(
            client=client,
            amount=amount,
            submitted_by=request.user,
            approved=False,
            interest_rate=37.8,
            exempt_day=exempt_day,
        )

        start_date = timezone.now().date() + timedelta(days=1)
        repayment_days = []
        count = 0
        day_offset = 0

        while count < 26:
            day = start_date + timedelta(days=day_offset)
            # Skip exempt day (Saturday or Sunday)
            if day.strftime('%A').upper() != exempt_day.upper():
                repayment_days.append(LoanRepaymentDay(
                    loan=loan,
                    date=day,
                    day_number=count + 1  # Day number from 1 to 26
                ))
                count += 1
            day_offset += 1

        LoanRepaymentDay.objects.bulk_create(repayment_days)

        messages.success(request, f"Loan for {client.full_name} submitted successfully.")
    except Exception as e:
        messages.error(request, f"Loan application failed: {str(e)}")

    return redirect("view_approved_clients")


@require_GET
def get_client_details(request, client_id):
    try:
        client = Client.objects.get(pk=client_id)
    except Client.DoesNotExist:
        raise Http404("Client not found")

    client_data = {
        "id": client.id,
        "full_name": client.full_name,
        "nrc": client.nrc,
        "phone_number": client.phone_number,
        "passport_photo": client.passport_photo.url if client.passport_photo else "",
        "signature": client.signature.url if client.signature else "",
        "business_name": client.business_name,
        "address": client.address,
        "marital_status": client.marital_status,
        "relationship_with_witness": client.relationship_with_witness,
        "surety_name": client.surety_name,
        "surety_value": client.surety_value,
        "surety_make": client.surety_make,
        "witness_name": client.witness_name,
        "witness_nrc": client.witness_nrc,
        "witness_phone": client.witness_phone,
    }

    return JsonResponse(client_data)


@require_http_methods(["GET", "POST"])
def edit_client_api(request, client_id):
    try:
        client = Client.objects.get(pk=client_id)
    except Client.DoesNotExist:
        return JsonResponse({'error': 'Client not found'}, status=404)

    if request.method == 'GET':
        # Return client data as JSON for pre-filling the modal
        data = {
            'full_name': client.full_name,
            'nrc': client.nrc,
            'phone_number': client.phone_number,
            'business_name': client.business_name,
            'address': client.address,
            'marital_status': client.marital_status,
            'relationship_with_witness': client.relationship_with_witness,
            'surety_name': client.surety_name,
            'surety_value': client.surety_value,
            'surety_make': client.surety_make,
            'witness_name': client.witness_name,
            'witness_nrc': client.witness_nrc,
            'witness_phone': client.witness_phone,
        }
        return JsonResponse(data)

    elif request.method == 'POST':
        form = ClientEditForm(request.POST, request.FILES, instance=client)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'errors': form.errors}, status=400)





@csrf_exempt
@require_http_methods(["POST"])
@login_required
def update_client(request):
    client_id = request.POST.get("client_id")
    try:
        client = Client.objects.get(id=client_id)
    except Client.DoesNotExist:
        return JsonResponse({"error": "Client not found."}, status=404)

    client.full_name = request.POST.get("full_name", client.full_name)
    client.nrc = request.POST.get("nrc", client.nrc)
    client.phone_number = request.POST.get("phone_number", client.phone_number)
    client.business_name = request.POST.get("business_name", client.business_name)
    client.address = request.POST.get("address", client.address)
    client.marital_status = request.POST.get("marital_status", client.marital_status)
    client.relationship_with_witness = request.POST.get("relationship_with_witness", client.relationship_with_witness)

    client.surety_name = request.POST.get("surety_name", client.surety_name)
    surety_value = request.POST.get("surety_value")
    if surety_value:
        try:
            client.surety_value = float(surety_value)
        except ValueError:
            pass
    client.surety_make = request.POST.get("surety_make", client.surety_make)

    client.witness_name = request.POST.get("witness_name", client.witness_name)
    client.witness_nrc = request.POST.get("witness_nrc", client.witness_nrc)
    client.witness_phone = request.POST.get("witness_phone", client.witness_phone)

    if request.FILES.get("passport_photo"):
        client.passport_photo = request.FILES["passport_photo"]

    if request.FILES.get("signature"):
        client.signature = request.FILES["signature"]

    client.save()

    return redirect('view_all_clients')


@csrf_exempt
def delete_client(request, client_id):
    if request.method == "DELETE":
        client = get_object_or_404(Client, id=client_id)
        client.delete()
        return JsonResponse({"success": True})
    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)


@login_required
def get_repayment_days(request, loan_id):
    repayment_days = LoanRepaymentDay.objects.filter(loan_id=loan_id).order_by('date')
    today = timezone.now().date()

    data = []

    for day in repayment_days:
        amount_due = day.amount_due or Decimal('0.00')
        amount_paid = day.amount_paid or Decimal('0.00')
        carried = day.balance_carried_forward or Decimal('0.00')

        data.append({
            'id': day.id,
            'date': day.date.strftime('%Y-%m-%d'),
            'is_paid': day.is_paid,
            'is_today': day.date == today,
            'amount_due': str(round(amount_due, 2)),
            'amount_paid': str(round(amount_paid, 2)) if day.is_paid else None,
            'balance_carried_forward': str(round(carried, 2)) if carried else None,
        })

    return JsonResponse(data, safe=False)

from decimal import Decimal, InvalidOperation
import json
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.db import transaction
from django.utils import timezone
from .models import LoanRepaymentDay

@require_POST
@login_required
@csrf_exempt
@require_POST
def mark_repayment_day_paid(request, day_id):
    try:
        repayment_day = get_object_or_404(LoanRepaymentDay, id=day_id)

        if repayment_day.is_paid:
            return JsonResponse(
                {'success': False, 'error': 'This day has already been marked as paid.'},
                status=400
            )

        # Parse payment amount from request

            return JsonResponse({'success': False, 'error': 'This day has already been marked as paid.'}, status=400)

        try:
            data = json.loads(request.body)
            raw_amount = data.get('amount_paid')
            amount_paid = Decimal(str(raw_amount)) if raw_amount else repayment_day.amount_due
        except (json.JSONDecodeError, InvalidOperation):

            return JsonResponse(
                {'success': False, 'error': 'Invalid payment amount.'},
                status=400
            )

        daily_payment = repayment_day.loan.daily_payment
        user = request.user if request.user.is_authenticated else None

        with transaction.atomic():
            # Calculate the effective due amount (daily payment + any carried shortfall)
            effective_due = daily_payment + (repayment_day.balance_carried_forward or Decimal('0.00'))
            
            # Mark current day as paid with exact amount
            repayment_day.amount_paid = amount_paid
            repayment_day.is_paid = True
            repayment_day.marked_by = user
            repayment_day.marked_at = timezone.now()
            
            # Reset the carried balance for this day since we're processing it now
            repayment_day.balance_carried_forward = Decimal('0.00')
            repayment_day.save()

            # Handle underpayment (shortfall)
            if amount_paid < effective_due:
                shortfall = effective_due - amount_paid
                # Apply shortfall to next unpaid day only
                next_unpaid_day = repayment_day.loan.repayment_days.filter(
                    date__gt=repayment_day.date,
                    is_paid=False
                ).order_by('date').first()
                
                if next_unpaid_day:
                    next_unpaid_day.balance_carried_forward = shortfall
                    next_unpaid_day.save()

            # Handle advance payment (excess) - only if payment exceeds effective due
            elif amount_paid > effective_due:
                excess = amount_paid - effective_due
                remaining_excess = excess
                
                # Get all future unpaid days in order
                future_days = repayment_day.loan.repayment_days.filter(
                    date__gt=repayment_day.date,
                    is_paid=False
                ).order_by('date')
                
                for future_day in future_days:
                    if remaining_excess <= 0:
                        break
                    
                    if remaining_excess >= daily_payment:
                        # Mark the full day as paid
                        future_day.amount_paid = daily_payment
                        future_day.is_paid = True
                        future_day.marked_by = user
                        future_day.marked_at = timezone.now()
                        remaining_excess -= daily_payment
                    else:
                        # Apply partial payment to next day only
                        future_day.amount_paid = remaining_excess
                        future_day.is_paid = False  # Not fully paid
                        # Set the carried forward balance to be negative (indicating an advance)
                        future_day.balance_carried_forward = -(daily_payment - remaining_excess)
                        remaining_excess = Decimal('0.00')
                    
                    future_day.save()

            # Check if all days are now paid
            if not repayment_day.loan.repayment_days.filter(is_paid=False).exists():
                repayment_day.loan.approved = True
                repayment_day.loan.save()

        return JsonResponse({
            'success': True,
            'message': f"Payment of ZMW {amount_paid:.2f} processed.",

            return JsonResponse({'success': False, 'error': 'Invalid payment amount.'}, status=400)

        # Mark current day as paid
        repayment_day.amount_paid = amount_paid
        repayment_day.is_paid = True
        repayment_day.marked_at = timezone.now()
        repayment_day.marked_by = request.user if request.user.is_authenticated else None
        repayment_day.save()

        delta = amount_paid - (repayment_day.amount_due + (repayment_day.balance_carried_forward or Decimal('0.00')))

        future_days = repayment_day.loan.repayment_days.filter(
            is_paid=False,
            date__gt=repayment_day.date
        ).order_by('date')

        if delta != 0 and future_days.exists():
            next_day = future_days.first()

            if next_day.balance_carried_forward is None:
                next_day.balance_carried_forward = Decimal("0.00")

            if delta < 0:
                next_day.balance_carried_forward += abs(delta)
            else:
                next_day.balance_carried_forward -= delta

            next_day.save()

        return JsonResponse({
            'success': True,
            'message': f"Payment of ZMW {amount_paid:.2f} recorded for {repayment_day.date}.",
            'adjustment': f"{'Shortfall' if delta < 0 else 'Excess'} of ZMW {abs(delta):.2f} {'carried to' if delta < 0 else 'deducted from'} next day." if delta != 0 else "Exact amount paid.",
            'date': repayment_day.date.strftime('%Y-%m-%d'),

            'amount_paid': float(amount_paid),
        })

    except LoanRepaymentDay.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Repayment day not found.'}, status=404)


@login_required
def ceo_dashboard(request):
    total_clients = Client.objects.filter(approved=True).count()
    pending_loans = Loan.objects.filter(approved=False).count()
    pending_clients = Client.objects.filter(approved=False).count()

    approved_loans = Loan.objects.filter(approved=True)
    total_expected = Decimal('0.00')

    for loan in approved_loans:
        amount_with_interest = loan.total_with_interest
        paid_amount = loan.repayment_days.filter(is_paid=True).aggregate(
            total=Sum('amount_paid')
        )['total'] or Decimal('0.00')

        balance = amount_with_interest - paid_amount
        total_expected += balance

    active_loans = [
        loan for loan in approved_loans
        if loan.repayment_days.filter(is_paid=True).count() < 26
    ]

    context = {
        'total_clients': total_clients,
        'pending_loans': pending_loans,
        'pending_clients': pending_clients,
        'total_expected': round(total_expected, 2),
        'active_loans': len(active_loans),
    }

    return render(request, 'dashboard/ceo_dashboard.html', context)


@require_POST
def approve_client(request, client_id):
    if not request.user.is_authenticated or request.user.role != 'CEO':
        return render(request, '403.html')  # or redirect('login') if preferred

    client = get_object_or_404(Client, id=client_id)

    if client.approved:
        messages.info(request, "Client is already approved.")
    else:
        client.approved = True
        client.save()
        messages.success(request, f"{client.full_name} has been approved successfully.")

    return redirect('view_all_clients')

@require_POST
@login_required
def reject_client(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    client.rejected = True
    client.approved = False
    client.save()
    messages.warning(request, f"Client {client.full_name} has been rejected.")
    return redirect('view_all_clients')


@login_required
def loan_officer_edit_profile(request):
    user = request.user
    if request.method == 'POST':
        form = LoanOfficerEditProfileForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save(commit=False)
            password = form.cleaned_data.get('password')
            if password:
                user.password = make_password(password)
            user.save()
            return redirect('officer_dashboard')  # Or wherever you'd like
    else:
        form = LoanOfficerEditProfileForm(instance=user)
    return render(request, 'loan_officer/edit_profile.html', {'form': form})

@login_required
def search_clients_by_nrc(request):
    query = request.GET.get("nrc", "").strip()
    clients = Client.objects.filter(nrc__icontains=query).order_by('-created_at') if query else Client.objects.all().order_by('-created_at')

    data = [{
        "id": c.id,
        "full_name": c.full_name,
        "nrc": c.nrc,
        "phone_number": c.phone_number,
        "created_at": c.created_at.strftime('%b %d, %Y'),
        "approved": c.approved
    } for c in clients]

    return JsonResponse({
        "clients": data,
        "csrf_token": get_token(request)
    })


@require_POST
def update_payment(request):
    try:
        payment_id = request.POST.get("payment_id")
        corrected_amount = Decimal(request.POST.get("corrected_amount", "0"))
        note = request.POST.get("edit_note", "").strip()

        repayment = get_object_or_404(LoanRepaymentDay, id=payment_id)

        if corrected_amount <= 0:
            messages.error(request, "Corrected amount must be greater than 0.")
            return redirect("ceo_view_all_payments")

        repayment.corrected_amount = corrected_amount
        repayment.edit_note = note
        repayment.edited_by = request.user
        repayment.edited_at = timezone.now()
        repayment.save()

        messages.success(request, f"Payment correction applied successfully.")
    except Exception as e:
        messages.error(request, f"Failed to update payment: {str(e)}")

    return redirect("ceo_view_all_payments")


@login_required
def create_loan_officer(request):

    if request.method == 'POST':
        form = LoanOfficerCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, " Loan officer account created successfully.")
            return redirect('create_loan_officer')
    else:
        form = LoanOfficerCreationForm()

    return render(request, 'ceo/create_loan_officer.html', {'form': form})
